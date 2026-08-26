"""Excel 计算书渲染（模板驱动）：迹树+结果 → 数据注入模板（零写死逻辑）。

输入:  TraceTree + PlantResult + 计算书模板 .xlsx（v1 调用方传入路径；
       正式模板归 data/templates 录入批 UF-16）
输出:  .xlsx 计算书（模板只做展示，全部数值由本文件写入）
"""

# ══════════════════════════════════════════════════════════════════
# 规格说明（M1b 实现，简报 D2 裁决 2026-08-25；镜像测试 tests/trace/
#   test_calcbook.py）
#
# 【公开接口】
#   class InvalidTemplateError(Exception)（GR-11 族新类，本文件定义）
#   TEMPLATE_REGISTRY: Final = ()   v1 空（正式模板归 data/templates
#       录入批 UF-16——登记后键→模板文件与占位符映射在此扩展）
#   render_calcbook(trace: TraceTree, result: PlantResult,
#                   template: Path, out: Path) -> Path
#
# 【行为规格】
#   R1 模板与数据分离：模板是版本化 .xlsx；本文件只做"占位符 → 值"
#       注入，写盘代码出现业务拼接 = 评审拒绝（占位符机制本身即分离）。
#   R2 模板禁公式（§11 R12）：加载模板后遍历全部工作表全部单元格，
#       data_type=="f"（含公式）→ InvalidTemplateError（消息含工作表
#       名与单元格坐标，GR-09）——计算单一事实源在 Python。
#   R3 数值按字段 ID 注入，占位符语法本简报冻结（v1 最小取值域）：
#       {{trace[i].<field>}}      —— TraceNode 六字段（i=迹序号）
#       {{trace[i].inputs.<symbol>}} —— 输入快照符号值
#       {{summary.<key>}}          —— 汇总平键（点式扁平键
#                                    f"{condition_key}.{字段ID}"，UF-42 同款）
#       未知占位符（不匹配语法/索引越界/键不存在）= InvalidTemplateError
#       （消息含原占位符与单元格坐标，GR-09）。
#   R4 字节确定性：保存经 ZipInfo 缺省时间戳重写 zip 条目（openpyxl
#       save 默认携带落盘时刻，双渲染字节不同——重写后双渲染字节相同）。
#
# 【数值纪律】本文件不在魔法数字白名单——零数值字面量（ZipInfo 缺省
#   date_time 即 zip 纪元，无字面量）。
#
# 【测试要求】最小模板渲染无 {{}} 残留、含公式模板拒、双渲染字节同、
#   未知占位符拒。
#
# 【参照】重写计划 §2/§6.5/§11 R12；简报 M1b D2；data/templates/README.md
# ══════════════════════════════════════════════════════════════════

from __future__ import annotations

import re
from collections.abc import Mapping
from io import BytesIO
from pathlib import Path
from types import MappingProxyType
from typing import Final
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook

from waterprint.contracts.result_schema import PlantResult
from waterprint.trace.collector import TraceTree

__all__ = ["TEMPLATE_REGISTRY", "InvalidTemplateError", "render_calcbook"]

# UF-16 收口（DRAFT 批 2026-08-26）：正式模板键→文件名映射
# （data/templates 1.0.0；占位符语法=本文件 _SUBST_PATTERN 冻结面）。
TEMPLATE_REGISTRY: Final[Mapping[str, str]] = MappingProxyType({
    "calcbook_unit": "calcbook_unit.xlsx",
    "calcbook_plant": "calcbook_plant.xlsx",
})

# 占位符语法（M1b D2 冻结）：{{trace[i].<field>}} / {{trace[i].inputs.<symbol>}
# / {{summary.<key>}}——本正则是语法的唯一裁定面（未匹配即未知占位符拒）。
_SUBST_PATTERN: re.Pattern[str] = re.compile(r"\{\{([^{}]+)\}\}")
_TRACE_FIELD: re.Pattern[str] = re.compile(
    r"trace\[(?P<idx>\d+)\]\.(?P<field>\w+)(?:\.(?P<sym>\w+))?\Z"
)
_SUMMARY_KEY: re.Pattern[str] = re.compile(r"summary\.\S+\Z")
_TRACE_FIELDS: Final[frozenset[str]] = frozenset(
    {"formula_id", "inputs", "output", "norm_ref", "unit_id", "condition_key"}
)
# summary 平键展开（R3）：{f"{condition_key}.{字段ID}": value}
_Summary = dict[str, float]


class InvalidTemplateError(Exception):
    """计算书模板非法（含公式单元格/未知占位符）——GR-11 族（M1b D2）。"""


def _summary_index(result: PlantResult) -> _Summary:
    """PlantResult.summary 两层映射 → 点式扁平键索引（{{summary.<key>}} 值域）。"""
    flat: _Summary = {}
    for condition_key, fields in result.summary.items():
        for field_id, value in fields.items():
            flat[f"{condition_key}.{field_id}"] = value
    return flat


def _resolve(token: str, where: str, trace: TraceTree, summary: _Summary) -> object:
    """单占位符 → 值；不可解析/越界/缺键 → InvalidTemplateError。

    六族拒消息一律携带单元格坐标 where（sheet!coordinate，GR-09——
    R1-a 二审 I-1 修复：where 原为死参数，现拼入全部未知占位符族）。
    """
    trace_match = _TRACE_FIELD.fullmatch(token)
    if trace_match is not None:
        index, field, symbol = (
            trace_match["idx"],
            trace_match["field"],
            trace_match["sym"],
        )
        if int(index) >= len(trace):
            raise InvalidTemplateError(
                f"占位符迹序号越界：{token!r}（迹长 {len(trace)}，位于 {where}）"
            )
        if symbol is not None:
            if field != "inputs" or symbol not in trace[int(index)].inputs:
                raise InvalidTemplateError(
                    f"占位符引用未绑定符号：{token!r}"
                    f"（迹节点 {index} 的 inputs 无 {symbol!r}，位于 {where}）"
                )
            return trace[int(index)].inputs[symbol]
        if field == "inputs":
            raise InvalidTemplateError(
                f"占位符语法不完整（inputs 须带 .<symbol>）：{token!r}"
                f"（位于 {where}）"
            )
        if field not in _TRACE_FIELDS:
            raise InvalidTemplateError(f"未知迹字段：{token!r}（位于 {where}）")
        return getattr(trace[int(index)], field)
    if _SUMMARY_KEY.fullmatch(token) is not None:
        key = token.removeprefix("summary.")
        if key not in summary:
            raise InvalidTemplateError(
                f"占位符引用未登记汇总键：{token!r}（位于 {where}）"
            )
        return summary[key]
    raise InvalidTemplateError(f"未知占位符语法：{token!r}（位于 {where}）")


def _render_cell(value: object, where: str, trace: TraceTree,
                  summary: _Summary) -> object:
    """单单元格占位符注入：非字符串原样返回；字符串内全部占位符替换。"""
    if not isinstance(value, str) or "{{" not in value:
        return value
    tokens = _SUBST_PATTERN.findall(value)
    outcomes = {token: _resolve(token, where, trace, summary) for token in tokens}
    if len(tokens) == 1 and value.replace(" ", "") == "{{" + tokens[0] + "}}":
        outcome = outcomes[tokens[0]]
        if isinstance(outcome, int | float | str):
            return outcome  # 整格单占位符：保持数值类型（Excel 数值格）
    rendered = value
    for token, outcome in outcomes.items():
        rendered = rendered.replace("{{" + token + "}}", str(outcome))
    return rendered


def _check_no_formulas(workbook: Workbook) -> None:
    """R2 模板禁公式：遍历全部工作表单元格，data_type=='f' 即拒。"""
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise InvalidTemplateError(
                        f"模板含公式单元格：{sheet.title}!{cell.coordinate}"
                        "（§11 R12——计算单一事实源在 Python，模板只做展示）"
                    )


def _deterministic_save(workbook: Workbook, out: Path) -> None:
    """R4 字节确定性保存：先入内存，再以 ZipInfo 缺省时间戳重写 zip 条目。"""
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    with ZipFile(buffer) as source, ZipFile(out, "w", ZIP_DEFLATED) as target:
        for name in source.namelist():
            entry = ZipInfo(name)  # 缺省 date_time=zip 纪元——确定性锚点
            entry.compress_type = source.getinfo(name).compress_type
            target.writestr(entry, source.read(name))


def render_calcbook(
    trace: TraceTree, result: PlantResult, template: Path, out: Path
) -> Path:
    """计算书渲染正门：模板禁公式检查 → 占位符注入 → 确定性保存（返回 out）。"""
    workbook = load_workbook(template)
    _check_no_formulas(workbook)
    summary = _summary_index(result)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = _render_cell(  # type: ignore[assignment]
                    cell.value, f"{sheet.title}!{cell.coordinate}", trace, summary
                )
    _deterministic_save(workbook, out)
    return out
