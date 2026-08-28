"""管网 Excel 读写：管段模型进、设计结果 sheet 出（M3 Excel 闭环）。

输入:  .xlsx 管网表（模板驱动：data/templates 管网模板）
输出:  PipeSegment 序列（读） / 带结果 sheet 的 .xlsx（写）

NET2 实装注记（2026-08-28，段二批）：
- 模板=data/templates/network_pipes.xlsx v1.0.0（sheet "pipes" 首行 7 列
  表头；第 2 行版本注记——segment_id 格以"模板版本"三字起始者识别为
  注记行跳过，数据自第 3 行起）；列位映射按表头名取列索引，禁硬编码
  列字母（R1——模板本体即列位）。
- 行数/大小上限经 assumptions network.excel.* 两键（§18 zip 炸弹护栏，
  NET2 批登记）；只读模式 openpyxl read_only 解析（R2）。
- 必填/类型/标高倒置校验一次性全报（R3）：NetworkExcelError 聚合
  带行号错误清单后抛出，禁止逐行崩溃。
- 结果 sheet 名 "results"：只写数值与文本零公式（R4）；幂等重写
  （存在即删重建，不重复叠加）；openpyxl 落盘天然 UTF-8（R5）；
  路径须绝对且无 ".." 分量（§18 同 dxf_writer 规则）。
- pipe_type 列 ∈ {concrete, plastic}（模板注记行明文集合——糙率键名
  口径）：read 校验合法性，水力糙率由调用方经 DesignOptions.roughness
  全局传入（规格头 PipeSegment 字段冻结不含管材，记档实现报告）。
"""

# ══════════════════════════════════════════════════════════════════
# 规格说明（骨架冻结；镜像测试 tests/network/test_excel_io.py）
#
# 【公开接口】
#   read_network_excel(path: Path) -> tuple[PipeSegment, ...]
#   write_result_sheet(path: Path, design: NetworkDesign) -> Path
#       原文件追加/更新结果 sheet（幂等：重写同 sheet 不重复）
#
# 【行为规格】
#   R1 模板驱动（§2 Excel 行）：读取按 data/templates 管网模板的
#      列位映射（模板版本号写入文件）；列位变更走模板版本化，
#      代码按模板描述取列，禁止硬编码列字母。
#   R2 安全面（§18 Excel zip 炸弹行）：大小上限（Settings 配置）、
#      行数上限、只读模式解析（openpyxl read_only）；超限抛领域异常。
#   R3 校验前置：必填列缺失/类型错误/标高倒置 → 带行号的错误清单
#      （一次性全报，禁止逐行崩溃）；读入结果进 solver 前已合法。
#   R4 输出无公式（§11 R12）：结果 sheet 只写数值与文本（计算全部
#      在 Python 完成），Excel 打开仅作展示。
#   R5 编码：openpyxl 读写显式 UTF-8 语义；文件路径限制在配置
#      工作目录内（§18 路径安全，同 dxf_writer 规则）。
#
# 【测试要求】模板往返（读→写→重读一致）、错误清单带行号、
#   行数/大小上限触发、幂等重写。
#
# 【参照】重写计划 §2 Excel 行/§11 R12/§18；data/templates/README.md
# ══════════════════════════════════════════════════════════════════

from __future__ import annotations

from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Final

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from waterprint.network.solver import (
    NetworkDesign,
    PipeSegment,
)
from waterprint.registry.assumptions import assumption

__all__ = ["NetworkExcelError", "read_network_excel", "write_result_sheet"]


class NetworkExcelError(Exception):
    """管网 Excel 读/写非法（上限防弹/校验清单/路径越界）——领域异常。"""


# 模板列名（data/templates/network_pipes.xlsx v1.0.0 表头逐字——列位映射
# 即模板本体，R1；列位变更走模板版本化）。
_COLUMNS: Final[tuple[str, ...]] = (
    "segment_id",
    "design_flow",
    "length",
    "ground_start",
    "ground_end",
    "upstream_invert",
    "pipe_type",
)
# 管材合法集（模板注记行明文 {concrete, plastic}——network.roughness.* 键名）。
_PIPE_TYPES: Final[frozenset[str]] = frozenset({"concrete", "plastic"})
# 注记行识别前缀（模板第 2 行"模板版本 …"形态——v1 固定，出处即模板本体）。
_NOTE_PREFIX: Final[str] = "模板版本"
# 结果 sheet 名（幂等重写目标——R4）。
_RESULT_SHEET: Final[str] = "results"
# 结果表头（segment_id + 设计结果六字段 + 满流参照，零公式纯展示）。
_RESULT_COLUMNS: Final[tuple[str, ...]] = (
    "segment_id",
    "diameter_m",
    "slope",
    "velocity_m_s",
    "depth_ratio",
    "invert_start_m",
    "invert_end_m",
    "v_full_m_s",
    "q_full_m3_s",
)
# 数值单元格形态（int/float 非 bool——GR-02 口径）。
type _CellValue = object


def _guard_size(path: Path) -> None:
    """R2 大小上限（§18 zip 炸弹护栏——assumptions network.excel.*）。"""
    limit = assumption("network.excel.max_file_bytes", {})
    size = path.stat().st_size
    if size > limit:
        raise NetworkExcelError(
            f"文件大小超上限：{path.name} {size} 字节 > {limit:.0f} 字节"
            "（§18 Excel zip 炸弹护栏，assumptions network.excel.max_file_bytes）"
        )


def _guard_rows(row_count: int) -> None:
    """R2 行数上限（read_only 模式 max_row 口径）。"""
    limit = assumption("network.excel.max_rows", {})
    if row_count > limit:
        raise NetworkExcelError(
            f"行数超上限：{row_count} > {limit:.0f}"
            "（§18 Excel 护栏，assumptions network.excel.max_rows）"
        )


def _header_index(header: Sequence[_CellValue]) -> dict[str, int]:
    """表头行 → {列名: 列索引}；缺任一模板列 = 领域异常（R1 模板驱动）。"""
    index = {
        str(cell).strip(): position for position, cell in enumerate(header) if cell is not None
    }
    missing = [name for name in _COLUMNS if name not in index]
    if missing:
        raise NetworkExcelError(
            f"表头缺模板列：{missing}（期望 {_COLUMNS}——模板 v1.0.0 列位，"
            "列位变更走模板版本化，R1）"
        )
    return index


def _numeric(cell: _CellValue) -> float | None:
    """数值单元格解析（int/float 非 bool 归一 float；空→None）。"""
    if cell is None or (isinstance(cell, str) and not cell.strip()):
        return None
    if isinstance(cell, bool) or not isinstance(cell, int | float):
        return None
    return float(cell)


def _validate_row(
    values: Mapping[str, _CellValue], row_number: int, is_first: bool
) -> tuple[list[str], PipeSegment | None]:
    """单数据行校验（R3）：返回 (错误清单, 解析段|None)。

    标高倒置（ground_start≤ground_end）在该层拒——读入结果进 solver
    前已合法；upstream_invert 首段必填、后续段留空=承接上段末管底
    （模板注记口径）。
    """
    errors: list[str] = []
    segment_id = values["segment_id"]
    if not isinstance(segment_id, str) or not segment_id.strip():
        errors.append(f"第 {row_number} 行 segment_id 必须为非空文本")
    pipe_type = values["pipe_type"]
    pipe_type_text = pipe_type.strip() if isinstance(pipe_type, str) else None
    if pipe_type_text not in _PIPE_TYPES:
        errors.append(
            f"第 {row_number} 行 pipe_type 非法：{pipe_type!r}"
            f"（合法 {sorted(_PIPE_TYPES)}——network.roughness.* 键名）"
        )
    flow_value = _numeric(values["design_flow"])
    length_value = _numeric(values["length"])
    ground_start = _numeric(values["ground_start"])
    ground_end = _numeric(values["ground_end"])
    for name, value in (
        ("design_flow", flow_value), ("length", length_value),
        ("ground_start", ground_start), ("ground_end", ground_end),
    ):
        if value is None:
            errors.append(f"第 {row_number} 行 {name} 必须为数值")
    upstream = _numeric(values["upstream_invert"])
    if is_first and upstream is None:
        errors.append("第 {row_number} 行 upstream_invert 首段必填（绝对标高 m）")
    if errors:
        return errors, None
    assert flow_value is not None  # errors 空 ⇔ 数值列全解析成功（类型面）
    assert length_value is not None
    assert ground_start is not None
    assert ground_end is not None
    if ground_start <= ground_end:
        return [
            f"第 {row_number} 行标高倒置：ground_start "
            f"{ground_start!r} ≤ ground_end {ground_end!r}"
            "（重力流倒坡，R3 校验前置）"
        ], None
    return [], PipeSegment(
        segment_id=str(segment_id).strip(),
        design_flow=flow_value,
        length=length_value,
        ground_start=ground_start,
        ground_end=ground_end,
        upstream_invert=upstream,
    )


def read_network_excel(path: Path) -> tuple[PipeSegment, ...]:
    """读管网表 → PipeSegment 序列（只读模式+双上限+校验清单一次性全报）。"""
    _guard_size(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["pipes"] if "pipes" in workbook.sheetnames else None
        if sheet is None:
            raise NetworkExcelError(
                "缺模板 sheet 'pipes'（data/templates/network_pipes.xlsx v1.0.0 形态）"
            )
        _guard_rows(sheet.max_row)
        rows = enumerate(sheet.iter_rows(values_only=True), start=1)
        _, header = next(rows, (0, ()))
        index = _header_index(header if isinstance(header, tuple) else ())
        errors: list[str] = []
        segments: list[PipeSegment] = []
        seen_data = False
        for row_number, row in rows:
            if not any(cell is not None for cell in row):
                continue
            values = {
                name: row[position] if position < len(row) else None
                for name, position in index.items()
            }
            first_cell = values["segment_id"]
            if isinstance(first_cell, str) and first_cell.startswith(_NOTE_PREFIX):
                continue
            row_errors, segment = _validate_row(values, row_number, not seen_data)
            errors.extend(row_errors)
            if segment is not None:
                segments.append(segment)
            seen_data = True
        if errors:
            raise NetworkExcelError("管网表校验失败（一次性全报）：\n" + "\n".join(errors))
        if not segments:
            raise NetworkExcelError("管网表无数据行（数据自第 3 行起录——模板 v1.0.0 注记口径）")
        return tuple(segments)
    finally:
        workbook.close()


def _validate_target_path(path: Path) -> Path:
    """R5 路径安全（§18 同 dxf_writer 规则）：绝对路径+拒 '..' 分量。"""
    if not path.is_absolute():
        raise NetworkExcelError(f"输出路径须为绝对路径：{str(path)!r}（§18 路径安全——SERVER 教训）")
    if ".." in path.parts:
        raise NetworkExcelError(f"输出路径含越界分量 '..'：{str(path)!r}（§18 路径安全）")
    return path


def write_result_sheet(
    path: Path,
    design: NetworkDesign,
    references: Mapping[str, tuple[float, float]] | None = None,
) -> Path:
    """幂等重写结果 sheet（R4/R5）：删旧建新、零公式、返回原路径。

    references：{segment_id: (v_full, q_full)} 满流参照补充列（可选——
    golden 附带锚口径，cli 传 None 时留空）。
    """
    target = _validate_target_path(path)
    _guard_size(target)
    workbook = load_workbook(target)
    try:
        if _RESULT_SHEET in workbook.sheetnames:
            del workbook[_RESULT_SHEET]
        sheet: Worksheet = workbook.create_sheet(_RESULT_SHEET)
        sheet.append(
            ("结果 sheet（network_pipes 模板 v1.0.0 口径；计算在 Python 完成，本表零公式，R4）",)
        )
        sheet.append(_RESULT_COLUMNS)
        extra = references or {}
        for result in design.results:
            v_full: float | str
            q_full: float | str
            if result.segment_id in extra:
                v_full, q_full = extra[result.segment_id]
            else:
                v_full = q_full = ""
            sheet.append(
                (
                    result.segment_id,
                    result.diameter,
                    result.slope,
                    result.velocity,
                    result.depth_ratio,
                    result.invert_start,
                    result.invert_end,
                    v_full,
                    q_full,
                )
            )
        for group in design.parallel:
            sheet.append(
                (
                    f"{group.segment_id}（并联×2，各输半量——用户可否决，R3）",
                    group.diameter,
                    "",
                    group.velocity,
                    group.depth_ratio,
                    "",
                    "",
                    "",
                    "",
                )
            )
        for failure in design.failures:
            sheet.append(
                (
                    f"{failure.segment_id}（无解段——违反约束清单，R5）",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    " | ".join(failure.reasons),
                    "",
                )
            )
        for warning in design.warnings:
            sheet.append((f"警示：{warning}",))
        workbook.save(target)
    finally:
        workbook.close()
    return target
