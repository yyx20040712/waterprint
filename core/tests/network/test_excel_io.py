"""excel_io 镜像测试：管网 Excel 读写（模板往返/错误带行号/上限防弹接线）。

输入:  waterprint.network.excel_io 公开符号
输出:  读写契约断言
"""

from __future__ import annotations

import importlib
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

_mod = importlib.import_module("waterprint.network.excel_io")
read_network_excel = getattr(_mod, "read_network_excel", None)
write_result_sheet = getattr(_mod, "write_result_sheet", None)

_solver = importlib.import_module("waterprint.network.solver")
DesignOptions = getattr(_solver, "DesignOptions")
PipeSegment = getattr(_solver, "PipeSegment")

pytestmark = pytest.mark.skipif(
    None in (read_network_excel, write_result_sheet),
    reason="实现未就绪：waterprint.network.excel_io（M3）",
)

_COLUMNS = (
    "segment_id",
    "design_flow",
    "length",
    "ground_start",
    "ground_end",
    "upstream_invert",
    "pipe_type",
)


def _write_template(path: Path) -> None:
    """照模板 v1.0.0 形态造最小管网表（表头+注记行+两数据行）。"""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "pipes"
    sheet.append(list(_COLUMNS))
    sheet.append(("模板版本 v1.0.0（测试 fixture——列位映射即模板本体）。",))
    sheet.append(("W1-W2", 0.05, 150.0, 52.0, 51.4, 48.0, "concrete"))
    sheet.append(("W2-W3", 0.12, 200.0, 51.4, 50.6, None, "concrete"))
    workbook.save(path)


def test_entrypoints_frozen() -> None:
    """入口冻结：read_network_excel(path) / write_result_sheet(path, design)。"""
    assert callable(read_network_excel)
    assert callable(write_result_sheet)


def test_template_roundtrip_wiring(tmp_path: Path) -> None:
    """R1 接线断言（NET2 填真实现）：读→写→重读一致（模板列位映射不漂移）。

    幂等面：二次 write 后 results sheet 恰一张（删旧建新不叠加）；
    结果 sheet 写入不扰 pipes sheet——重读管段与首次读入完全一致。
    """
    xlsx = tmp_path / "pipes.xlsx"
    _write_template(xlsx)
    segments = read_network_excel(xlsx)
    assert segments == (
        PipeSegment(
            segment_id="W1-W2",
            design_flow=0.05,
            length=150.0,
            ground_start=52.0,
            ground_end=51.4,
            upstream_invert=48.0,
        ),
        PipeSegment(
            segment_id="W2-W3",
            design_flow=0.12,
            length=200.0,
            ground_start=51.4,
            ground_end=50.6,
            upstream_invert=None,
        ),
    )
    options = DesignOptions(
        available_diameters=(0.4, 0.5),
        min_velocity=0.6,
        max_velocity=3.0,
        max_depth=6.0,
        fill_ratio_steps=((0.45, 0.65), (0.9, 0.70)),
        roughness=0.013,
    )
    design = _solver.design_pipes(segments, options)
    assert design.failures == ()
    out = write_result_sheet(xlsx, design)
    assert out == xlsx
    assert read_network_excel(xlsx) == segments
    workbook = load_workbook(xlsx, read_only=True)
    try:
        names = workbook.sheetnames
        rows = list(workbook["results"].iter_rows(values_only=True))
    finally:
        workbook.close()
    assert names.count("results") == 1
    assert any(row[0] == "W1-W2" for row in rows)
    write_result_sheet(xlsx, design)
    workbook = load_workbook(xlsx, read_only=True)
    try:
        assert workbook.sheetnames.count("results") == 1
    finally:
        workbook.close()
