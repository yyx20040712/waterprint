"""市政污水 golden 端到端（34,760 m³/d，一级 A；M2 验收）。

输入:  golden_data/municipal_34760/{input_project.json, expected_summary.json}
输出:  全流程与实跑期望值对照断言（5 工况终水逐项+design 档主尺寸+
       计算书导出+2+k 工况索引——GOLDEN 批激活，2026-08-26；D10
       2026-08-28：summary 真值断言+正式 calcbook_plant 模板渲染收口；
       GOLDEN2 2026-08-28：污泥链扩图 12→19 节点+m3_deferred 真值对照
       ——D3 summary 多汇点未触发（拓扑执行序 bashi 仍居末位汇点））
"""

# ══════════════════════════════════════════════════════════════════
# 规格：M2 验收 = 13 单元全线 + 汇流 + 高程 + 枚举诊断端到端出全套
# 计算书；期望值来源 docs/norms 手算对照与旧系统结果（差异逐条解释，
# 由领域专家签字后录入 golden_data——实现者不得自编）。
# ══════════════════════════════════════════════════════════════════

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pytest

pytestmark = [
    pytest.mark.golden,
    pytest.mark.skipif(
        not (Path(__file__).parent / "golden_data" / "municipal_34760" / "expected_summary.json").is_file(),
        reason="golden 数据未整理（M0 §9.4：34,760 m³/d 案例由领域专家录入）",
    ),
]

_REPO_DATA = Path(__file__).resolve().parents[3] / "data" / "coefficients"
_REPO_TEMPLATES = Path(__file__).resolve().parents[3] / "data" / "templates"
_REPO_PRICES = Path(__file__).resolve().parents[3] / "data" / "unit_prices"
_TERMINAL = "municipal_bashi_jiliangcao"


def _m3_real_values(plant: Any, expected: dict[str, Any]) -> None:
    """M3 面真值（GOLDEN2 2026-08-28）：概算总数+全厂总泥量实跑对照。

    estimate_total=全图（19 节点）design 档工程量→概算——app 未接 cost
    属既定架构（result_schema"愿景未落"注记），测试直调 cost 三正门
    （tests/cost/test_estimate.py 先例）；grand_total 逐级自洽
    （subtotal+reserve_subtotal+Σtax=grand_total）先证后对照。
    total_sludge=hebing ds_total（干基 kg/d 主口径）；湿基 q_total 以
    design_dims["sludge_hebing"]["q_total"] 锚承载双断言。"""
    from waterprint.cost.estimate import build_estimate, load_fee_rules
    from waterprint.cost.prices import load_prices
    from waterprint.cost.takeoff import takeoff_quantities

    book = load_prices(_REPO_PRICES)
    fees = load_fee_rules(_REPO_PRICES / "field_mapping.yaml", book)
    items = takeoff_quantities(plant, "design", price_book=book)
    sheet = build_estimate(items, book, fees)
    assert (sheet.subtotal + sheet.reserve_subtotal
            + sum(line.amount for line in sheet.tax)) == sheet.grand_total
    m3 = expected["m3_deferred"]
    for key in ("estimate_total", "total_sludge"):
        assert set(m3[key]) == {"value", "source", "abs", "rel"}, key
    estimate = m3["estimate_total"]
    assert sheet.grand_total == pytest.approx(
        estimate["value"], rel=estimate["rel"], abs=estimate["abs"]
    ), "m3_deferred.estimate_total（19 节点 design 档 grand_total）"
    sludge = m3["total_sludge"]
    hebing = plant.conditions["design"]["sludge_hebing"].dims
    assert hebing["ds_total"] == pytest.approx(
        sludge["value"], rel=sludge["rel"], abs=sludge["abs"]
    ), "m3_deferred.total_sludge（hebing ds_total 干基 kg/d 主口径）"
    wet = expected["design_dims"]["sludge_hebing"]["q_total"]
    assert hebing["q_total"] == pytest.approx(
        wet["value"], rel=wet["rel"], abs=wet["abs"]
    ), "hebing q_total（湿基 m³/d——m3 双断言第二锚）"


def _calcbook_min_template_renders(
    plant: Any, expected: dict[str, Any], tmp_path: Path
) -> None:
    """④ 计算书导出面（最小模板）：trace/summary 两族占位符各≥1，全展开零残留。

    D10 2026-08-28：summary 真值=run_full_calc app 层注入（executor
    summary={} 占位经 app 层 replace 回填——_summary_of 纯投影），本用例
    自造 replace workaround 已移除，直接消费 plant.summary 真值。
    """
    from openpyxl import Workbook, load_workbook

    from waterprint.app import export_artifact

    summary_view = plant.summary["design"]  # 真值消费（replace workaround 移除）
    assert set(summary_view) == set(expected["effluent"]["design"])  # 键集钳制
    for indicator, item in expected["effluent"]["design"].items():
        assert summary_view[indicator] == pytest.approx(
            item["value"], rel=item["rel"], abs=item["abs"]
        ), f"summary design.{indicator}"
    template = tmp_path / "calcbook_tpl.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "{{trace[0].formula_id}}"
    workbook.active["B1"] = "BOD5={{summary.design.BOD5}}"
    workbook.save(template)
    payload = export_artifact(
        "calcbook", plant, template, tmp_path / "calcbook_out.xlsx"
    )
    assert payload  # 产出非空（渲染成功即无 InvalidTemplateError）
    sheet = load_workbook(tmp_path / "calcbook_out.xlsx").active
    assert sheet["A1"].value == plant.trace[0].formula_id  # trace 占位符→实跑迹真值
    assert sheet["B1"].value == f"BOD5={summary_view['BOD5']}"  # summary 占位符→实跑值
    assert "{{" not in f"{sheet['A1'].value}{sheet['B1'].value}"  # 零残留


def _calcbook_official_template_renders(
    plant: Any, expected: dict[str, Any], tmp_path: Path
) -> None:
    """④ 续：正式模板真值渲染（D10 收口，冻结裁决项④）——data/templates。

    模板路径=对账 fixture parents 法（同 _REPO_DATA 口径解析 data/
    templates）；经 export_artifact("calcbook") 渲染后断言：全表占位符
    零残留（"{{" 计数 0）+ B3:B8 六指标值==expected design 逐项
    （1e-12 双容差）——templates manifest 1.0.0 平键集复核的机器锚定。"""
    from openpyxl import load_workbook

    from waterprint.app import export_artifact

    template = _REPO_TEMPLATES / "calcbook_plant.xlsx"
    out = tmp_path / "calcbook_plant_out.xlsx"
    assert export_artifact("calcbook", plant, template, out)  # 渲染成功零拒
    sheet = load_workbook(out).active
    assert sheet.title == "全厂汇总"  # 唯一表（templates 1.0.0 契约）
    rendered = "".join(
        f"{cell.value}" for row in sheet.iter_rows() for cell in row
    )
    assert "{{" not in rendered  # 六占位符全展开零残留
    for row, indicator in enumerate(
        ("BOD5", "CODCR", "SS", "NH3N", "TN", "TP"), start=3
    ):
        item = expected["effluent"]["design"][indicator]
        assert sheet[f"B{row}"].value == pytest.approx(
            item["value"], rel=item["rel"], abs=item["abs"]
        ), f"正式模板 {indicator}"


def test_municipal_golden_end_to_end(golden_data_dir: Path, tmp_path: Path) -> None:
    """端到端：run_full_calc 5 工况终水逐项+主尺寸+计算书导出对照期望（双容差禁放宽）。

    占位实质化（GOLDEN 批总授权"占位实质化"先例，2026-08-26）：原占位
    raise 的语义全部承载——①正门跑 input_project（5 工况全）②逐工况逐项
    对照 expected（rel/abs 1e-12 双容差按 expected 内标注）③主尺寸逐项
    对照 ④计算书导出（最小模板+正式 calcbook_plant 占位符全展开真值）
    ⑤工况集 2+k 索引断言（iter_all 键集恰等 5 预期）。
    """
    from waterprint.app import load_project, run_full_calc
    from waterprint.contracts.condition import ConditionSet, build_condition_set
    from waterprint.contracts.run_env import RunEnv
    from waterprint.registry import load_coefficients
    from waterprint.registry.assumptions import DEFAULT_ASSUMPTIONS

    case_dir = golden_data_dir / "municipal_34760"
    project = load_project(case_dir / "input_project.json")
    expected = json.loads(
        (case_dir / "expected_summary.json").read_text(encoding="utf-8")
    )

    # ⑤ 工况集 2+k 索引（ADR-007：基线两档+每受检单元一条，键集恰等预期）
    conditions = build_condition_set(expected["checked_units"])
    keys = [ConditionSet.key(c) for c in conditions.iter_all()]
    assert keys == expected["condition_keys"]
    assert len(keys) == 2 + len(expected["checked_units"]) == 5

    # ① 正门实跑（env 口径=expected.generated 实录：server 正门版本串）
    lib = load_coefficients(_REPO_DATA)
    env = RunEnv(
        engine_version=expected["generated"]["engine_version"],
        data_version=expected["generated"]["data_version"],
        assumptions={entry.key: entry.default for entry in DEFAULT_ASSUMPTIONS},
        coefficients=lib,
        price_book={},
        trace_sink=None,
        engine_params={},
    )
    plant = run_full_calc(project, conditions, env).plant
    assert set(plant.conditions) == set(keys)  # 全 5 工况各出整图结果
    assert plant.repro.design_hash == project.metadata.content_hash  # 结果绑定输入
    assert set(plant.summary) == set(keys)  # D10 全工况注入口径（逐工况映射）

    # ② 逐工况逐项终水对照（双容差按 expected 内标注——不放宽）
    # 键集钳制（GOLDEN R1-2）：expected 工况块恰等 5 工况键集，防删块静默绿
    assert set(expected["effluent"]) == set(keys)
    for condition_key, fields in expected["effluent"].items():
        snapshot = plant.conditions[condition_key][_TERMINAL]
        for indicator, item in fields.items():
            actual = snapshot.outqualities[f"{_TERMINAL}.out.{indicator}"]
            assert actual == pytest.approx(
                item["value"], rel=item["rel"], abs=item["abs"]
            ), f"终水 {condition_key}.{indicator}"

    # ③ design 档主尺寸逐项对照（每 unit 主控项，容差同上）
    # 单元覆盖钳制（GOLDEN R1-2 + GOLDEN2 扩面）：design_dims 恰覆盖
    # 十九节点减 inlet（污泥链 7 单元 2026-08-28 入锚——只增不改）
    assert set(expected["design_dims"]) == set(project.design.nodes) - {"inlet"}
    for unit_id, fields in expected["design_dims"].items():
        dims = plant.conditions["design"][unit_id].dims
        for field, item in fields.items():
            assert dims[field] == pytest.approx(
                item["value"], rel=item["rel"], abs=item["abs"]
            ), f"主尺寸 {unit_id}.{field}"

    # M3 面真值（GOLDEN2 2026-08-28）：概算总数/全厂总泥量实跑对照
    # （结构={value, source, abs, rel}——与 design_dims 条目同形态；
    # grand_total 逐级自洽+hebing ds_total/q_total 双断言，禁字串回退）
    _m3_real_values(plant, expected)

    _calcbook_min_template_renders(plant, expected, tmp_path)
    _calcbook_official_template_renders(plant, expected, tmp_path)
