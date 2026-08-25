"""calcbook 镜像测试：计算书渲染（占位符零残留/模板禁公式/确定性接线）。

输入:  waterprint.trace.calcbook 公开符号
输出:  渲染契约断言（§11 R12——模板只展示，计算在 Python）
"""

from __future__ import annotations

import importlib
from pathlib import Path

import pytest

_mod = importlib.import_module("waterprint.trace.calcbook")
render_calcbook = getattr(_mod, "render_calcbook", None)
InvalidTemplateError = getattr(_mod, "InvalidTemplateError", None)

pytestmark = pytest.mark.skipif(
    render_calcbook is None,
    reason="实现未就绪：waterprint.trace.calcbook（M1）",
)


def _trace() -> tuple[object, ...]:
    """最小迹树：单 TraceNode（六字段全量）。"""
    from waterprint.contracts.result_schema import TraceNode

    return (
        TraceNode(
            formula_id="M1B-CALCBOOK-F1",
            inputs={"a": 1.0, "b": 2.0},
            output=3.0,
            norm_ref="测试条文 M1b-calcbook",
            unit_id="m1b_probe_unit",
            condition_key="design",
        ),
    )


def _result() -> object:
    """最小 PlantResult（summary 平键取值域：design.total_sludge）。"""
    from types import MappingProxyType

    from waterprint.contracts.result_schema import PlantResult, ReproTriple

    return PlantResult(
        conditions=MappingProxyType({}),
        summary=MappingProxyType(
            {"design": MappingProxyType({"total_sludge": 667.4})}
        ),
        trace=(),  # type: ignore[arg-type]
        repro=ReproTriple(
            design_hash="m1b", engine_version="m1b", data_version="m1b"
        ),
    )


def _template(path: Path, a1: object, b1: object = "静态文本") -> Path:
    """测试夹具自造最小 xlsx 模板（tmp_path 内 openpyxl 现场生成，不经 data/）。"""
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.active["A1"] = a1
    workbook.active["B1"] = b1
    workbook.save(path)
    return path


def test_entrypoint_frozen() -> None:
    """入口冻结：render_calcbook(trace, result, template, out)。"""
    assert callable(render_calcbook)


def test_no_marker_residue_wiring(tmp_path: Path) -> None:
    """R1 接线断言：渲染产物中模板标记零残留。"""
    from openpyxl import load_workbook

    template = _template(
        tmp_path / "tpl.xlsx",
        "{{trace[0].formula_id}}",
        "输出={{trace[0].output}}；DS={{summary.design.total_sludge}}",
    )
    out = render_calcbook(_trace(), _result(), template, tmp_path / "out.xlsx")  # type: ignore[misc]
    assert isinstance(out, Path)
    sheet = load_workbook(out).active
    assert sheet["A1"].value == "M1B-CALCBOOK-F1"
    assert "{{" not in str(sheet["B1"].value)  # 最小模板渲染后无 {{field_id}} 残留
    assert "3.0" in str(sheet["B1"].value)
    assert "667.4" in str(sheet["B1"].value)


def test_formula_template_rejected_wiring(tmp_path: Path) -> None:
    """R2 接线断言：含 Excel 公式的模板加载即失败。"""
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.active["A1"] = "=SUM(B1:B2)"  # 构造含公式单元格模板断言拒绝
    template = tmp_path / "formula_tpl.xlsx"
    workbook.save(template)
    with pytest.raises(InvalidTemplateError, match="A1"):
        render_calcbook(_trace(), _result(), template, tmp_path / "out.xlsx")  # type: ignore[misc]
