"""exports 路由镜像测试：导出端点（stale 守门、文件名安全）。

输入:  waterprint_server.routers.exports 公开符号
输出:  路由契约断言
"""

from __future__ import annotations

import asyncio
import importlib

import pytest
from fastapi import status

_mod = importlib.import_module("waterprint_server.routers.exports")
router = getattr(_mod, "router")

pytestmark = [
    pytest.mark.skipif(
        router is None,
        reason="实现未就绪：waterprint_server.routers.exports（服务层 M2/M3）",
    ),
]

_EXPECTED = {
    ("post", "/api/exports/calcbook"),
    ("post", "/api/exports/audit"),
    ("post", "/api/exports/dxf"),
    ("post", "/api/exports/estimate"),
    ("post", "/api/exports/ifc"),
    ("get", "/api/exports"),
}


async def _project_with_result(client, extra_units=()):  # type: ignore[no-untyped-def]
    """建项目并跑 calc 至 done（FE9 R2：extra_units 链式追加处理单元节点）。"""
    nodes: dict[str, object] = {
        "inlet": {
            "kind": "municipal_input",
            "q_avg_daily": 34760.7 / 86400,
            "kz": 1.4,
            "CODCR": 400.0,
            "BOD5": 200.0,
            "SS": 250.0,
            "NH3N": 26.0,
            "TN": 43.0,
            "TP": 6.5,
        },
        "municipal_cass": {},
    }
    for unit_id in extra_units:
        nodes[unit_id] = {}
    chain = ["inlet", *extra_units, "municipal_cass"]
    edges = [
        {
            "src": {"unit_id": chain[i], "port_id": "out"},
            "dst": {"unit_id": chain[i + 1], "port_id": "in"},
        }
        for i in range(len(chain) - 1)
    ]
    payload = {
        "project": {
            "format_version": "1.0",
            "design": {"nodes": nodes, "edges": edges},
            "view": {},
            "metadata": {
                "format_version": "1.0",
                "content_hash": "0",
                "engine_version": "0",
                "data_version": "0",
            },
        }
    }
    created = await client.post("/api/projects", json=payload)
    project_id = created.json()["project_id"]
    task_id = (await client.post(
        "/api/calc/run", json={"project_id": project_id, "conditions": []}
    )).json()["task_id"]
    for _ in range(300):
        body = (await client.get(f"/api/calc/tasks/{task_id}")).json()
        if body.get("state") in {"done", "failed"}:
            break
        await asyncio.sleep(0.1)
    assert body["state"] == "done"
    return project_id, task_id  # type: ignore[no-any-return]


def test_router_exposes_five_endpoints_wiring() -> None:
    """端点集 == 规格六件（calcbook/audit/dxf/estimate/ifc/列表——SC1 ifc 增）。"""
    observed = {
        (method.lower(), route.path) for route in router.routes for method in route.methods
    }  # type: ignore[union-attr]
    assert observed >= _EXPECTED and len(observed) == len(_EXPECTED)  # 恰六件无漂移


@pytest.mark.anyio
async def test_stale_result_returns_409_with_context_wiring(client) -> None:  # type: ignore[no-untyped-def]
    """R1 接线断言：结果集三元组过期且未 force → 409 附输入版本信息与摘要值。"""
    project_id, task_id = await _project_with_result(client)
    tasks = (await client.get(f"/api/calc/tasks/{task_id}")).json()
    result_digest = tasks["result"]["design_hash"]  # 结果集三元组摘要真值
    project = (await client.get(f"/api/projects/{project_id}")).json()
    project["design"]["assumption_overrides"] = {"safety.superheight": 0.3}
    saved = await client.put(f"/api/projects/{project_id}", json=project)
    assert saved.status_code == status.HTTP_200_OK and saved.json()["design_changed"] is True
    stale = await client.post("/api/exports/calcbook", json={"project_id": project_id})
    assert stale.status_code == status.HTTP_409_CONFLICT
    detail = str(stale.json()["detail"])
    assert "输入版本" in detail  # 附输入版本信息
    assert result_digest[:6] in detail  # 摘要值锁定（AU-6/R1-4②——非仅关键词）
    forced = await client.post(
        "/api/exports/calcbook", json={"project_id": project_id}, params={"force": "true"}
    )
    assert forced.status_code == status.HTTP_200_OK  # force=1 导出旧结果（文件流）
    metas = await client.get("/api/exports", params={"project_id": project_id})
    assert any(meta["stale_labeled"] for meta in metas.json())  # 旧三元组显式标注


@pytest.mark.anyio
async def test_calcbook_product_content_readback_wiring(client) -> None:  # type: ignore[no-untyped-def]
    """AUDIT2 I-9：calcbook xlsx 产物内容级读回（生成物面虚锚收口）。

    L1 抽查实录：conftest 占位符模板 {{trace[0].unit_id}} 造好后全库
    零 load_workbook 消费——渲染错误/空产物不会红（FE8「包内不可证」
    族）。本用例打开产物断：非空 sheet+占位符已被替换（trace 单元
    id 真值出现——模板键零残留）。
    """
    from io import BytesIO

    from openpyxl import load_workbook

    project_id, _task_id = await _project_with_result(client)
    fresh = await client.post("/api/exports/calcbook", json={"project_id": project_id})
    assert fresh.status_code == status.HTTP_200_OK
    workbook = load_workbook(BytesIO(fresh.content))
    assert workbook.sheetnames, "产物应有工作表"
    sheet = workbook[workbook.sheetnames[0]]
    cells = [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value]
    joined = "\n".join(cells)
    assert "{{" not in joined, "模板占位符应全部被渲染替换"
    assert "municipal_cass" in joined, "trace 单元 id 真值应入产物（占位符替换实证）"


@pytest.mark.anyio
async def test_dxf_product_flow_wiring(client) -> None:  # type: ignore[no-untyped-def]
    """FE9：dxf 单产物正向全链（D2 模板闸收窄+D3 options 透传+D4 后缀映射）。

    断：POST dxf{unit_id,design}→200 文件流（DXF R2018 魔面 AC1032 头+
    SECTION 实体节字节；Content-Disposition 与产物名 .dxf 后缀）；
    GET /api/exports 元数据行 kind=dxf（三元组摘要入边车）；bare POST
    dxf=全厂总图（M5 接线后）——200+.dxf 后缀+全厂总图 meta.title 行
    （$PROJECTNAME 落 DXF 头；无-unit 命名分量与单元图互异零覆盖）。
    """
    project_id, _task_id = await _project_with_result(client)
    dxf = await client.post(
        "/api/exports/dxf",
        json={
            "project_id": project_id,
            "condition_key": "design",
            "options": {"unit_id": "municipal_cass"},
        },
    )
    assert dxf.status_code == status.HTTP_200_OK
    assert b"AC1032" in dxf.content[:512]  # DXF R2018 头魔面（write_dxf 落盘）
    assert b"SECTION" in dxf.content  # 实体节标记（plan+section 图元真出图）
    disposition = str(dxf.headers.get("content-disposition", ""))
    # R4（DS-09）：解析文件名后 endswith 真锁后缀边界（非子串包含）
    disposition_name = disposition.rsplit('filename="', maxsplit=1)[-1].rstrip('"')
    assert disposition_name.endswith(".dxf")  # D4：kind 后缀映射
    metas = await client.get("/api/exports", params={"project_id": project_id})
    rows = [meta for meta in metas.json() if meta["kind"] == "dxf"]
    assert len(rows) == 1, "本用例恰导出一次（前置行数断言——非行序依赖）"
    assert rows[0]["file_name"].endswith(".dxf")  # D4 注册表口径同款
    assert rows[0]["stale_labeled"] is False  # 新鲜导出无 stale 标注
    bare = await client.post("/api/exports/dxf", json={"project_id": project_id})
    assert bare.status_code == status.HTTP_200_OK  # bare POST=全厂总图（M5 接线）
    assert b"AC1032" in bare.content[:512]  # 总图同为 DXF R2018
    assert "全厂总图".encode() in bare.content  # meta.title 行（$PROJECTNAME 头）
    bare_disposition = str(bare.headers.get("content-disposition", ""))
    bare_name = bare_disposition.rsplit('filename="', maxsplit=1)[-1].rstrip('"')
    assert bare_name.endswith(".dxf")  # M5：总图产物 .dxf 后缀
    metas = await client.get("/api/exports", params={"project_id": project_id})
    rows = [meta for meta in metas.json() if meta["kind"] == "dxf"]
    assert len(rows) == 2  # 单元图+总图两行（无-unit 命名互异零覆盖）
    assert len({meta["file_name"] for meta in rows}) == 2


@pytest.mark.anyio
async def test_dxf_stale_force_flow_wiring(client) -> None:  # type: ignore[no-untyped-def]
    """FE9 R2（DS-02）：dxf 409/force 端到端（calcbook stale fixture 同模式）。

    断：PUT 改 design 不重算→POST dxf 恰 409 StaleExportError（附输入
    版本+结果集摘要值）→force=1 重发→200 DXF 文件流+列表 dxf 行恰一行
    stale_labeled=true（旧三元组显式标注——R1 守门服务端契约回归锁）。
    """
    project_id, task_id = await _project_with_result(client)
    tasks = (await client.get(f"/api/calc/tasks/{task_id}")).json()
    result_digest = tasks["result"]["design_hash"]  # 结果集三元组摘要真值
    project = (await client.get(f"/api/projects/{project_id}")).json()
    project["design"]["assumption_overrides"] = {"safety.superheight": 0.3}
    saved = await client.put(f"/api/projects/{project_id}", json=project)
    assert saved.status_code == status.HTTP_200_OK and saved.json()["design_changed"] is True
    body = {
        "project_id": project_id,
        "condition_key": "design",
        "options": {"unit_id": "municipal_cass"},
    }
    stale = await client.post("/api/exports/dxf", json=body)
    assert stale.status_code == status.HTTP_409_CONFLICT
    assert stale.json()["error_type"] == "StaleExportError"
    detail = str(stale.json()["detail"])
    assert "输入版本" in detail  # 附输入版本信息
    assert result_digest[:6] in detail  # 摘要值锁定（AU-6/R1-4②——非仅关键词）
    forced = await client.post(
        "/api/exports/dxf", json=body, params={"force": "true"}
    )
    assert forced.status_code == status.HTTP_200_OK
    assert b"AC1032" in forced.content[:512]  # force 旧结果照常真出图
    metas = await client.get("/api/exports", params={"project_id": project_id})
    rows = [meta for meta in metas.json() if meta["kind"] == "dxf"]
    assert len(rows) == 1
    assert rows[0]["stale_labeled"] is True  # 旧三元组显式标注


@pytest.mark.anyio
async def test_dxf_multi_unit_distinct_names_wiring(
    client, test_settings
) -> None:  # type: ignore[no-untyped-def]
    """FE9 R2（DS-01 回归锁）：同项目同工况多单元导出文件名互异。

    断：cass+chenshachi 两单元先后导出→两文件名互异且各含 unit 分量
    （-dxf-municipal_{cass,chenshachi}-design-）→列表恰两行→exports_dir
    恰两个 .dxf（R1 前同名 os.replace 覆盖=首产物静默丢失缺陷收口）。
    """
    import os

    project_id, _task_id = await _project_with_result(
        client, extra_units=("municipal_chenshachi",)
    )
    body = {
        "project_id": project_id,
        "condition_key": "design",
    }
    first = await client.post(
        "/api/exports/dxf",
        json={**body, "options": {"unit_id": "municipal_cass"}},
    )
    assert first.status_code == status.HTTP_200_OK
    second = await client.post(
        "/api/exports/dxf",
        json={**body, "options": {"unit_id": "municipal_chenshachi"}},
    )
    assert second.status_code == status.HTTP_200_OK
    metas = await client.get("/api/exports", params={"project_id": project_id})
    rows = [meta for meta in metas.json() if meta["kind"] == "dxf"]
    assert len(rows) == 2, "两单元产物均应注册（同名覆盖时仅剩一行）"
    names = {meta["file_name"] for meta in rows}
    assert len(names) == 2  # 文件名互异
    assert any("-dxf-municipal_cass-design-" in name for name in names)
    assert any("-dxf-municipal_chenshachi-design-" in name for name in names)
    dxf_files = sorted(
        name for name in os.listdir(test_settings.exports_dir) if name.endswith(".dxf")
    )
    assert len(dxf_files) == 2  # exports_dir 双 dxf（无同名覆盖）


@pytest.mark.anyio
async def test_traversal_export_rejected_wiring(client, test_settings) -> None:  # type: ignore[no-untyped-def]
    """AU-1/R1-1 路由面：condition_key/items kind 穿越 → 422 且产物零新增落盘。"""
    import os

    project_id, _task_id = await _project_with_result(client)
    exports_dir = test_settings.exports_dir
    before_listing = sorted(os.listdir(exports_dir))
    before_files = sorted(str(p) for p in exports_dir.parent.parent.rglob("*"))
    shallow = await client.post(
        "/api/exports/calcbook",
        json={"project_id": project_id, "condition_key": "a/../../evil"},
    )
    assert shallow.status_code == status.HTTP_422_UNPROCESSABLE_CONTENT  # 越界即拒
    deep = await client.post(
        "/api/exports/calcbook",
        json={"project_id": project_id, "condition_key": "b/" + "../" * 10 + "deep"},
    )
    assert deep.status_code == status.HTTP_422_UNPROCESSABLE_CONTENT
    batch = await client.post(
        "/api/exports/calcbook",
        json={
            "project_id": project_id,
            "options": {"items": [{"kind": "calcbook/../../evil", "condition_key": ""}]},
        },
    )
    assert batch.status_code == status.HTTP_422_UNPROCESSABLE_CONTENT
    assert sorted(os.listdir(exports_dir)) == before_listing  # exports_dir 零新增
    assert sorted(str(p) for p in exports_dir.parent.parent.rglob("*")) == before_files


@pytest.mark.anyio
async def test_ifc_export_flow_wiring(client, test_settings) -> None:  # type: ignore[no-untyped-def]
    """SC1 D7：ifc 单产物正向全链——200 文件流+.ifc 后缀+IFC 魔面非空。

    断：POST ifc{design}→200（core build_scene→build_ifc→write_ifc 链——
    body ISO-10303-21 STEP 头魔面非空）；Content-Disposition 文件名
    .ifc 后缀（_KIND_SUFFIXES 五键映射）；GET /api/exports 元数据行
    kind=ifc 且 .ifc 后缀；exports_dir 落盘真件（下载流=落盘件读取）。
    """
    import os

    project_id, _task_id = await _project_with_result(client)
    model = await client.post(
        "/api/exports/ifc",
        json={"project_id": project_id, "condition_key": "design"},
    )
    assert model.status_code == status.HTTP_200_OK
    assert model.content, "ifc 产物禁空字节（UF-33）"
    assert b"ISO-10303-21" in model.content[:512]  # IFC STEP 物理文件头魔面
    disposition = str(model.headers.get("content-disposition", ""))
    disposition_name = disposition.rsplit('filename="', maxsplit=1)[-1].rstrip('"')
    assert disposition_name.endswith(".ifc")  # D7：kind 后缀映射
    metas = await client.get("/api/exports", params={"project_id": project_id})
    rows = [meta for meta in metas.json() if meta["kind"] == "ifc"]
    assert len(rows) == 1
    assert rows[0]["file_name"].endswith(".ifc")
    assert (test_settings.exports_dir / rows[0]["file_name"]).is_file()  # 落盘真件


@pytest.mark.anyio
async def test_ifc_traversal_project_id_rejected_wiring(  # type: ignore[no-untyped-def]
    client, test_settings
) -> None:
    """SC1 D7 质量门条款 4（路径拼接 DoD）：project_id 穿越 → 4xx 且导出目录快照零新增。

    浅（a/../../evil）深（../×10）两例——穿越 project_id 先死于 404
    ExportSourceNotFoundError（无此项目=无结果集，§18 路径安全族——
    穿越串永不进文件名/路径分量）。
    """
    import os

    _project_id, _task_id = await _project_with_result(client)
    exports_dir = test_settings.exports_dir
    before_listing = sorted(os.listdir(exports_dir))
    before_files = sorted(str(p) for p in exports_dir.parent.parent.rglob("*"))
    shallow = await client.post(
        "/api/exports/ifc",
        json={"project_id": "a/../../evil", "condition_key": "design"},
    )
    assert shallow.status_code == status.HTTP_404_NOT_FOUND  # 4xx（穿越≠结果集）
    deep = await client.post(
        "/api/exports/ifc",
        json={"project_id": "b/" + "../" * 10 + "deep", "condition_key": "design"},
    )
    assert deep.status_code == status.HTTP_404_NOT_FOUND
    assert sorted(os.listdir(exports_dir)) == before_listing  # exports_dir 零新增
    assert sorted(str(p) for p in exports_dir.parent.parent.rglob("*")) == before_files


async def _wait_task_terminal(client, task_id: str) -> dict:  # type: ignore[no-untyped-def]
    """轮询任务至终态（SVRB 批量转任务 E2E 消费面——复用 calc 查询面）。"""
    for _ in range(300):
        body = (await client.get(f"/api/calc/tasks/{task_id}")).json()
        if body.get("state") in {"done", "cancelled", "failed"}:
            return body  # type: ignore[no-any-return]
        await asyncio.sleep(0.1)
    pytest.fail(f"任务 {task_id} 300 轮询内未到终态（SVRB 批量 E2E）")


@pytest.mark.anyio
async def test_ifc_batch_no_unit_flow_wiring(client, test_settings) -> None:  # type: ignore[no-untyped-def]
    """SVRB D3：ifc 批量放行（原 422 拒绝面解除——project_path 通道等价透传）。

    无-unit ifc items×2（归一后全 None=批内一致过小闸）→200 句柄 JSON 含
    task_id→终态 done→exports_dir 两 .ifc+双边车（模型级命名无 unit 分量
    ——互异由 condition 分量保证）。
    """
    import os

    project_id, _task_id = await _project_with_result(client)
    batch = await client.post(
        "/api/exports/ifc",
        json={
            "project_id": project_id,
            "options": {"items": [
                {"kind": "ifc", "condition_key": "design"},
                {"kind": "ifc", "condition_key": "avg"},
            ]},
        },
    )
    assert batch.status_code == status.HTTP_200_OK  # 批量转任务句柄 JSON（原 422 面）
    task_id = str(batch.json()["task_id"])
    assert task_id  # 句柄含 task_id（webapp 任务态消费面）
    done = await _wait_task_terminal(client, task_id)
    assert done["state"] == "done" and len(done["result"]["files"]) == 2
    listing = sorted(os.listdir(test_settings.exports_dir))
    assert len([n for n in listing if n.endswith(".ifc")]) == 2  # 两模型落盘
    assert len([n for n in listing if n.endswith(".ifc.meta.json")]) == 2  # ifc 边车在


@pytest.mark.anyio
async def test_dxf_batch_no_unit_flow_wiring(client, test_settings) -> None:  # type: ignore[no-untyped-def]
    """SVRB D2：无-unit dxf 批量（全厂总图×2 工况）放行——worker site_design
    透传通道实证（原 422 拒绝面解除；总图内容=单产物 bare POST 同款）。"""
    import os

    project_id, _task_id = await _project_with_result(client)
    batch = await client.post(
        "/api/exports/dxf",
        json={
            "project_id": project_id,
            "options": {"items": [
                {"kind": "dxf", "condition_key": "design"},
                {"kind": "dxf", "condition_key": "avg"},
            ]},
        },
    )
    assert batch.status_code == status.HTTP_200_OK
    done = await _wait_task_terminal(client, str(batch.json()["task_id"]))
    assert done["state"] == "done" and len(done["result"]["files"]) == 2
    dxf_names = [
        n for n in sorted(os.listdir(test_settings.exports_dir)) if n.endswith(".dxf")
    ]
    assert len(dxf_names) == 2
    for name in dxf_names:  # 总图实证：site_design 透传→$PROJECTNAME 头（单产物同款）
        assert "全厂总图".encode() in (test_settings.exports_dir / name).read_bytes()


@pytest.mark.anyio
async def test_batch_missing_item_kind_normalizes_to_endpoint_wiring(
    client,
) -> None:
    """R2/G1-02+SVRB：per-item kind 缺省归一端点 kind——归一面放行闭合
    （原两族 422 拒绝面 SVRB 解除：dxf 归一→总图批量/ifc 归一→模型批量；
    R 轮 R1：补终态等待——200+task_id 仅证提交，done 才证归一面全链）。"""
    project_id, _task_id = await _project_with_result(client)
    for endpoint in ("dxf", "ifc"):
        batch = await client.post(
            f"/api/exports/{endpoint}",
            json={
                "project_id": project_id,
                "options": {"items": [
                    {"condition_key": "design"},
                    {"condition_key": "avg"},
                ]},
            },
        )
        assert batch.status_code == status.HTTP_200_OK  # 归一后放行转任务
        done = await _wait_task_terminal(client, str(batch.json()["task_id"]))
        assert done["state"] == "done"  # R1（R 轮）：终态实证


@pytest.mark.anyio
async def test_dirty_coord_grid_returns_422_and_no_artifacts(
    client, test_settings
) -> None:  # type: ignore[no-untyped-def]
    """ENG7 笔1：脏 site 几何（coord_grid=0.0）→ 422 InvalidSitePlanError 零落盘。

    触发链：POST /api/exports/dxf 不带 options.unit_id（全厂总图通道）→
    site_layout 坐标网间距闸（非有限/非正双拦——schema 对 coord_grid 无
    gt/finite 面，脏值经 PUT 放行存盘）。conditions 空路 API 不可常规
    构造（calc 恒产工况），core 单测 test_site_plan.py:463-482 已覆盖，
    本用例不重复——消费面行为断言锚 coord_grid 路（质量门条款 4：HTTP
    响应体实值，非名义表等值断言）。
    """
    import os

    project_id, _task_id = await _project_with_result(client)
    project = (await client.get(f"/api/projects/{project_id}")).json()
    project["design"]["site"]["options"]["coord_grid"] = 0.0  # 脏 site 几何（结构性非法）
    saved = await client.put(f"/api/projects/{project_id}", json=project)
    assert saved.status_code == status.HTTP_200_OK and saved.json()["design_changed"] is True
    rerun = await client.post(
        "/api/calc/run", json={"project_id": project_id, "conditions": []}
    )
    task_id = rerun.json()["task_id"]
    for _ in range(300):
        body = (await client.get(f"/api/calc/tasks/{task_id}")).json()
        if body.get("state") in {"done", "failed"}:
            break
        await asyncio.sleep(0.1)
    assert body["state"] == "done"  # 重算消 stale（导出消费新结果集）
    exports_dir = test_settings.exports_dir
    before_listing = sorted(os.listdir(exports_dir))
    resp = await client.post(
        "/api/exports/dxf",
        json={"project_id": project_id, "condition_key": "design"},  # 无 unit_id=总图通道
    )
    assert resp.status_code == status.HTTP_422_UNPROCESSABLE_CONTENT
    assert resp.json()["error_type"] == "InvalidSitePlanError"
    assert "坐标网间距非法" in resp.json()["detail"]
    assert sorted(os.listdir(exports_dir)) == before_listing  # 拒绝即零落盘
